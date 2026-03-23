"""
Monte Carlo Simulation for Bitcoin Mining Investment
=====================================================
Runs N simulated 60-month paths through the mining P&L model,
using historically calibrated distribution parameters for BTC price
and network difficulty.

Methodology:
  1. Calibrate: Extract monthly log-return distributions from historical
     data (mean, std dev, correlation) for pre-halving and post-halving regimes.
  2. Simulate: Generate N correlated random paths for BTC price and difficulty
     using Cholesky decomposition. Derive hashprice from the standard formula.
  3. Evaluate: Run each path through the mining P&L (revenue, costs, tax, cashflow)
     using fixed project parameters (capacity, PPA rate, miner specs, capex).
  4. Aggregate: Compute percentile distributions (P10/P25/P50/P75/P90) for
     IRR, payback period, net income, and cumulative cashflow.

Output:
  - Excel workbook with probability distributions and sample paths
  - Summary statistics for investment committee presentation

Usage:
  python monte_carlo_mining.py [--simulations 10000] [--months 60] [--output-dir ./output]

Dependencies:
  pip install numpy pandas openpyxl scipy numpy_financial

Built for the AGM Omerelu 10MW Bitcoin Mining Project.
"""

import argparse
import calendar
import copy
import os
import shutil
import warnings
from datetime import datetime, date

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# scipy.optimize no longer needed — using numpy polynomial root-finding for IRR

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Constants — fixed project parameters from Project 1 sheet
# ---------------------------------------------------------------------------
CAPACITY_MW = 10
CAPACITY_KW = CAPACITY_MW * 1000
EFFICIENCY_J_PER_TH = 21
FLEET_HASHRATE_THS = 420_042  # TH/s
UPTIME = 0.90
PPA_RATE = 0.0385  # $/kWh
OM_RATE = 0.004    # $/kWh
AXXELA_SHARE = 0.07
TAX_RATE = 0.30
DEPRECIATION_MONTHLY = 23_627.36
TOTAL_CAPEX = 2_332_023.24
STARTING_BTC_PRICE = 75_000.0
STARTING_DIFFICULTY = 145_000_000_000_000.0  # 145T

HISTORICAL_HALVING_DATE = datetime(2024, 4, 1)  # regime split for calibration (Apr 2024 halving)
SIM_START = datetime(2026, 4, 1)
PRE_HALVING_REWARD = 3.125
POST_HALVING_REWARD = 1.5625
HALVING_SIM_MONTH = 25  # month 25 = April 2028 (next halving during simulation)
DIFFICULTY_SHOCK = -0.10  # one-time -10% difficulty shock at halving (thinner margins → more exits than 2024)


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def days_in_month(year, month):
    """Return number of days in a given month."""
    return calendar.monthrange(year, month)[1]


def compute_monthly_irr(cashflows):
    """Compute monthly IRR via bisection, then annualize.

    Fast and robust for the 61-element cashflow vectors in this simulation.
    Annualized as: (1 + monthly_irr)^12 - 1

    Returns -1.0 (total loss) if the investment never breaks even and no
    IRR root can be found.
    """
    cfs = np.array(cashflows, dtype=np.float64)
    if len(cfs) < 2:
        return np.nan

    # If all cashflows are same sign, no IRR exists
    has_neg = np.any(cfs < 0)
    has_pos = np.any(cfs > 0)
    if not (has_neg and has_pos):
        # All negative = total loss
        if has_neg:
            return -1.0
        return np.nan

    t = np.arange(len(cfs), dtype=np.float64)

    def npv(r):
        with np.errstate(over="ignore", invalid="ignore"):
            return np.sum(cfs / (1.0 + r) ** t)

    # Check undiscounted sum: if sum < 0, investment loses money overall
    total = cfs.sum()

    # Bisection between lo and hi (monthly rate)
    lo, hi = -0.99, 2.0
    npv_lo = npv(lo)
    npv_hi = npv(hi)

    if np.isnan(npv_lo) or np.isnan(npv_hi):
        return -1.0 if total < 0 else np.nan

    if npv_lo * npv_hi > 0:
        # No sign change in this range
        for hi_try in [5.0, 10.0, 50.0]:
            npv_hi = npv(hi_try)
            if not np.isnan(npv_hi) and npv_lo * npv_hi <= 0:
                hi = hi_try
                break
        else:
            # No root found — return -1 if net loser, else NaN
            return -1.0 if total < 0 else np.nan

    # Bisect
    for _ in range(100):
        mid = (lo + hi) / 2.0
        npv_mid = npv(mid)
        if np.isnan(npv_mid):
            return -1.0 if total < 0 else np.nan
        if abs(npv_mid) < 1e-6:
            break
        if npv_lo * npv_mid < 0:
            hi = mid
        else:
            lo = mid
            npv_lo = npv_mid

    monthly_irr = (lo + hi) / 2.0
    annual_irr = (1.0 + monthly_irr) ** 12 - 1.0
    return annual_irr


def compute_moic(cashflows, capex):
    """MOIC = total cumulative payout / capex."""
    cfs = np.array(cashflows)
    # Sum all positive free cashflows (months 1+)
    inflows = cfs[1:][cfs[1:] > 0].sum()
    if capex == 0:
        return np.nan
    return inflows / capex


def compute_payback_month(cashflows):
    """First month where cumulative cashflow > 0 (0-indexed cashflows)."""
    cum = np.cumsum(cashflows)
    pos = np.where(cum > 0)[0]
    if len(pos) == 0:
        return np.nan
    return int(pos[0])  # month index (0 = capex month)


# ---------------------------------------------------------------------------
# Step 1 — Calibration
# ---------------------------------------------------------------------------
def calibrate(data_path):
    """Read historical data and compute parameters for both BASE and BULL scenarios.

    BASE CASE: Full historical regimes (pre-halving vs post-halving) as observed.
    BULL CASE: Rally-phase calibration with post-halving difficulty reduction.
    """
    print("=" * 60)
    print("STEP 1: CALIBRATION (BASE + BULL)")
    print("=" * 60)

    df = pd.read_excel(data_path, sheet_name="Daily Data")
    df.columns = df.columns.str.strip()

    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)

    print(f"  Historical data: {df['Date'].min().date()} to {df['Date'].max().date()}")
    print(f"  Total daily observations: {len(df)}")

    # Monthly averages
    df["YearMonth"] = df["Date"].dt.to_period("M")
    monthly = df.groupby("YearMonth").agg({
        "BTC Price (USD)": "mean",
        "Network Difficulty": "mean",
    }).reset_index()
    monthly.columns = ["YearMonth", "price", "difficulty"]
    monthly = monthly.sort_values("YearMonth").reset_index(drop=True)

    monthly["price_lr"] = np.log(monthly["price"] / monthly["price"].shift(1))
    monthly["diff_lr"] = np.log(monthly["difficulty"] / monthly["difficulty"].shift(1))
    monthly = monthly.dropna().reset_index(drop=True)
    monthly["date"] = monthly["YearMonth"].dt.to_timestamp()

    # Split points
    halving_ts = pd.Timestamp(HISTORICAL_HALVING_DATE)
    rally_start = pd.Timestamp("2024-04-01")
    rally_end = pd.Timestamp("2025-04-01")

    pre_full = monthly[monthly["date"] < halving_ts].copy()
    post_full = monthly[monthly["date"] >= halving_ts].copy()
    rally = monthly[(monthly["date"] >= rally_start) &
                    (monthly["date"] < rally_end)].copy()

    print(f"\n  Monthly observations: {len(monthly)}")
    print(f"  Pre-halving:   {len(pre_full)} months "
          f"({pre_full['date'].min().date()} – {pre_full['date'].max().date()})")
    print(f"  Post-halving:  {len(post_full)} months "
          f"({post_full['date'].min().date()} – {post_full['date'].max().date()})")
    print(f"  Rally phase:   {len(rally)} months "
          f"({rally['date'].min().date()} – {rally['date'].max().date()})")

    def compute_regime(subset):
        return {
            "mu_price": subset["price_lr"].mean(),
            "sigma_price": subset["price_lr"].std(ddof=1),
            "mu_diff": subset["diff_lr"].mean(),
            "sigma_diff": subset["diff_lr"].std(ddof=1),
            "correlation": subset[["price_lr", "diff_lr"]].corr().iloc[0, 1],
        }

    pre_params = compute_regime(pre_full)
    post_params = compute_regime(post_full)
    rally_params = compute_regime(rally)

    # ── BASE CASE: full historical regimes as-is ──
    base = {
        "pre_halving": pre_params,
        "post_halving": post_params,
    }

    # ── BULL CASE: rally-phase calibration + adjusted post-halving ──
    bull = {
        "pre_halving": rally_params.copy(),
        "post_halving": {
            "mu_price": rally_params["mu_price"],
            "sigma_price": rally_params["sigma_price"],
            "mu_diff": rally_params["mu_diff"] * 0.5,
            "sigma_diff": rally_params["sigma_diff"],
            "correlation": rally_params["correlation"] * 0.7,
        },
    }

    # Print both
    for scenario_name, params in [("BASE CASE", base), ("BULL CASE", bull)]:
        print(f"\n  ── {scenario_name} ──")
        for regime_key in ["pre_halving", "post_halving"]:
            p = params[regime_key]
            label = "Months 1–24" if regime_key == "pre_halving" else "Months 25–60"
            print(f"    {label}:")
            print(f"      Price  μ={p['mu_price']:+.6f}  σ={p['sigma_price']:.6f}"
                  f"  ({(np.exp(p['mu_price'] * 12) - 1) * 100:+.1f}%/yr)")
            print(f"      Diff   μ={p['mu_diff']:+.6f}  σ={p['sigma_diff']:.6f}"
                  f"  ({(np.exp(p['mu_diff'] * 12) - 1) * 100:+.1f}%/yr)")
            print(f"      Corr = {p['correlation']:.4f}")

    print(f"\n  Bull case rationale:")
    print(f"    - Rally-phase calibration (ETFs, institutional adoption = structural)")
    print(f"    - Q4-2025/Q1-2026 crash excluded (macro noise: tariffs, rate hikes)")
    print(f"    - Post-2028-halving: diff growth halved (miner capitulation)")
    print(f"    - Difficulty shock at halving: {DIFFICULTY_SHOCK:.0%}")
    print(f"    - Entry at ${STARTING_BTC_PRICE:,.0f} (35% below ATH)")

    return {"base": base, "bull": bull}


# ---------------------------------------------------------------------------
# Step 2 & 3 — Simulation + P&L
# ---------------------------------------------------------------------------
def run_simulations(params, n_sims, n_months):
    """Run Monte Carlo simulations and compute P&L for each path."""
    print("\n" + "=" * 60)
    print(f"STEP 2 & 3: SIMULATION ({n_sims:,} runs × {n_months} months)")
    print("=" * 60)

    np.random.seed(42)

    # Build Cholesky matrices for each regime
    chol = {}
    for regime in ["pre_halving", "post_halving"]:
        p = params[regime]
        cov = np.array([
            [p["sigma_price"] ** 2, p["correlation"] * p["sigma_price"] * p["sigma_diff"]],
            [p["correlation"] * p["sigma_price"] * p["sigma_diff"], p["sigma_diff"] ** 2],
        ])
        chol[regime] = np.linalg.cholesky(cov)

    # Build date/days-in-month schedule
    schedule = []
    y, m = SIM_START.year, SIM_START.month
    for i in range(n_months):
        dim = days_in_month(y, m)
        schedule.append((y, m, dim, date(y, m, 1)))
        m += 1
        if m > 12:
            m = 1
            y += 1

    # Pre-allocate result arrays
    all_irr = np.full(n_sims, np.nan)
    all_moic = np.full(n_sims, np.nan)
    all_payback = np.full(n_sims, np.nan)
    all_y1_net_income = np.full(n_sims, np.nan)
    all_5y_net_income = np.full(n_sims, np.nan)

    # Month-by-month cumulative cashflow for percentile curves (n_sims × n_months+1)
    all_cum_cf = np.zeros((n_sims, n_months + 1))  # month 0 = capex

    # Store 20 sample paths for detail output
    sample_indices = np.linspace(0, n_sims - 1, 20, dtype=int)
    sample_paths = {}

    report_interval = max(1, n_sims // 10)

    for sim in range(n_sims):
        if sim % report_interval == 0:
            print(f"  Simulation {sim:>6,} / {n_sims:,} ({100 * sim / n_sims:.0f}%)")

        price = STARTING_BTC_PRICE
        difficulty = STARTING_DIFFICULTY

        # Cashflows: month 0 = capex outflow, months 1..n = free cashflow
        cashflows = np.zeros(n_months + 1)
        cashflows[0] = -TOTAL_CAPEX

        monthly_net_income = np.zeros(n_months)

        path_data = [] if sim in sample_indices else None

        for month_idx in range(n_months):
            sim_month = month_idx + 1  # 1-indexed
            yr, mo, dim, dt = schedule[month_idx]

            # Determine regime and block reward
            if sim_month < HALVING_SIM_MONTH:
                regime = "pre_halving"
                block_reward = PRE_HALVING_REWARD
            else:
                regime = "post_halving"
                block_reward = POST_HALVING_REWARD

            # Generate correlated random draws
            p = params[regime]
            z = np.random.randn(2)
            corr_z = chol[regime] @ z

            lr_price = p["mu_price"] + corr_z[0]
            lr_diff = p["mu_diff"] + corr_z[1]

            # Apply halving difficulty shock at month 25
            if sim_month == HALVING_SIM_MONTH:
                lr_diff += np.log(1 + DIFFICULTY_SHOCK)

            # Update price and difficulty
            price = price * np.exp(lr_price)
            difficulty = difficulty * np.exp(lr_diff)
            difficulty = max(difficulty, 1e12)  # floor

            # Hashprice
            hashprice = price * block_reward * 86400 / (difficulty * 2 ** 32) * 1e15

            # Monthly BTC mined
            btc_mined = (FLEET_HASHRATE_THS * 1e12 * 86400 * block_reward *
                         UPTIME * dim / (difficulty * 2 ** 32))

            # Revenue
            revenue = btc_mined * price
            net_revenue = revenue * (1 - AXXELA_SHARE)

            # Costs
            hours = UPTIME * 24 * dim
            ppa_cost = CAPACITY_KW * hours * PPA_RATE
            om_cost = CAPACITY_KW * hours * OM_RATE

            # P&L with operational shutdown option:
            # If net revenue doesn't cover variable costs (PPA + O&M),
            # the mine idles that month — no revenue, no variable costs, FCF = 0.
            gross_profit = net_revenue - ppa_cost - om_cost
            if gross_profit < 0:
                # Shutdown: mine idles, no cash in or out
                ebitda = 0.0
                ebit = 0.0
                tax = 0.0
                net_income = 0.0
                free_cashflow = 0.0
                btc_mined = 0.0
                revenue = 0.0
            else:
                ebitda = gross_profit
                ebit = ebitda - DEPRECIATION_MONTHLY
                tax = ebit * TAX_RATE if ebit > 0 else 0.0
                net_income = ebit - tax
                free_cashflow = net_income + DEPRECIATION_MONTHLY

            cashflows[month_idx + 1] = free_cashflow
            monthly_net_income[month_idx] = net_income

            if path_data is not None:
                path_data.append({
                    "month": sim_month,
                    "date": dt,
                    "btc_price": price,
                    "difficulty": difficulty,
                    "hashprice": hashprice,
                    "btc_mined": btc_mined,
                    "revenue": revenue,
                    "net_income": net_income,
                    "free_cashflow": free_cashflow,
                })

        # Cumulative cashflow
        cum_cf = np.cumsum(cashflows)
        all_cum_cf[sim, :] = cum_cf

        # Metrics
        all_irr[sim] = compute_monthly_irr(cashflows)
        all_moic[sim] = compute_moic(cashflows, TOTAL_CAPEX)
        all_payback[sim] = compute_payback_month(cashflows)
        all_y1_net_income[sim] = monthly_net_income[:12].sum()
        all_5y_net_income[sim] = monthly_net_income.sum()

        if sim in sample_indices:
            sample_paths[sim] = {
                "path_data": path_data,
                "cum_cf": cum_cf.copy(),
            }

    print(f"  Simulation {n_sims:>6,} / {n_sims:,} (100%)")

    return {
        "irr": all_irr,
        "moic": all_moic,
        "payback": all_payback,
        "y1_net_income": all_y1_net_income,
        "y5_net_income": all_5y_net_income,
        "cum_cf": all_cum_cf,
        "sample_paths": sample_paths,
        "schedule": schedule,
    }


# ---------------------------------------------------------------------------
# Step 4 — Aggregation
# ---------------------------------------------------------------------------
def aggregate(results):
    """Compute percentiles and probability metrics."""
    print("\n" + "=" * 60)
    print("STEP 4: AGGREGATION")
    print("=" * 60)

    pcts = [10, 25, 50, 75, 90]
    stats = {}

    for key in ["irr", "moic", "payback", "y1_net_income", "y5_net_income"]:
        arr = results[key]
        valid = arr[~np.isnan(arr)]
        stats[key] = {f"P{p}": np.percentile(valid, p) if len(valid) > 0 else np.nan for p in pcts}
        stats[key]["mean"] = np.nanmean(arr)
        stats[key]["valid_count"] = len(valid)

    # Probability thresholds
    irr = results["irr"]
    valid_irr = irr[~np.isnan(irr)]
    n_valid = len(valid_irr)
    n_total = len(irr)

    probs = {}
    probs["irr_gt_0"] = np.sum(valid_irr > 0) / n_total if n_total > 0 else 0
    probs["irr_gt_10"] = np.sum(valid_irr > 0.10) / n_total if n_total > 0 else 0
    probs["irr_gt_20"] = np.sum(valid_irr > 0.20) / n_total if n_total > 0 else 0

    payback = results["payback"]
    valid_pb = payback[~np.isnan(payback)]
    probs["payback_24"] = np.sum(valid_pb <= 24) / n_total if n_total > 0 else 0
    probs["payback_36"] = np.sum(valid_pb <= 36) / n_total if n_total > 0 else 0
    probs["payback_48"] = np.sum(valid_pb <= 48) / n_total if n_total > 0 else 0

    # Probability of losing money = 5yr cumulative CF < 0
    final_cum = results["cum_cf"][:, -1]
    probs["lose_money"] = np.sum(final_cum < 0) / n_total if n_total > 0 else 0

    # Month-by-month percentile curves
    n_months_plus1 = results["cum_cf"].shape[1]
    cf_percentiles = {}
    for p in pcts:
        cf_percentiles[f"P{p}"] = np.percentile(results["cum_cf"], p, axis=0)

    # IRR histogram
    irr_valid = valid_irr[np.isfinite(valid_irr)]
    if len(irr_valid) > 0:
        # Clip extreme IRRs for histogram
        irr_clipped = np.clip(irr_valid, -1.0, 5.0)
        hist_counts, hist_edges = np.histogram(irr_clipped, bins=30)
    else:
        hist_counts, hist_edges = np.array([]), np.array([])

    # Print summary
    print("\n  IRR Percentiles:")
    for p in pcts:
        v = stats["irr"][f"P{p}"]
        print(f"    P{p}: {v:.1%}" if not np.isnan(v) else f"    P{p}: N/A")

    print(f"\n  MOIC Percentiles:")
    for p in pcts:
        v = stats["moic"][f"P{p}"]
        print(f"    P{p}: {v:.2f}x" if not np.isnan(v) else f"    P{p}: N/A")

    print(f"\n  Payback Month Percentiles:")
    for p in pcts:
        v = stats["payback"][f"P{p}"]
        print(f"    P{p}: {v:.0f}" if not np.isnan(v) else f"    P{p}: N/A")

    print(f"\n  Probability Thresholds:")
    print(f"    IRR > 0%:   {probs['irr_gt_0']:.1%}")
    print(f"    IRR > 10%:  {probs['irr_gt_10']:.1%}")
    print(f"    IRR > 20%:  {probs['irr_gt_20']:.1%}")
    print(f"    Payback ≤ 24m: {probs['payback_24']:.1%}")
    print(f"    Payback ≤ 36m: {probs['payback_36']:.1%}")
    print(f"    Payback ≤ 48m: {probs['payback_48']:.1%}")
    print(f"    Lose money: {probs['lose_money']:.1%}")

    print(f"\n  Year-1 Net Income — P50: ${stats['y1_net_income']['P50']:,.0f}")
    print(f"  5-Year Net Income — P50: ${stats['y5_net_income']['P50']:,.0f}")

    return {
        "stats": stats,
        "probs": probs,
        "cf_percentiles": cf_percentiles,
        "hist_counts": hist_counts,
        "hist_edges": hist_edges,
    }


# ---------------------------------------------------------------------------
# Step 5 — Write Excel
# ---------------------------------------------------------------------------
DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"

HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
BODY_FONT = Font(name="Calibri", size=10, color=BLACK)
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
)

FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_CURRENCY_DEC = '#,##0.00;(#,##0.00);"-"'
FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"
FMT_NUMBER = "#,##0"
FMT_NUMBER_DEC = "#,##0.00"
FMT_NUMBER_6 = "0.000000"
FMT_DATE = "MMM-YYYY"


def write_section_header(ws, row, title, max_col=8):
    """Write a dark blue section header row with 'x' in column A."""
    ws.cell(row=row, column=1, value="x").font = Font(name="Calibri", size=10, color=WHITE)
    ws.cell(row=row, column=2, value=title).font = HEADER_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="left")
    return row + 1


def write_sub_header(ws, row, headers, max_col=None):
    """Write a medium-blue sub-header row."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = SUB_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    mc = max_col or len(headers)
    for c in range(len(headers) + 1, mc + 1):
        ws.cell(row=row, column=c).fill = SUB_HEADER_FILL
    return row + 1


def write_kv_row(ws, row, label, value, fmt=None, alt=False):
    """Write a label-value row in columns B-C."""
    cell_b = ws.cell(row=row, column=2, value=label)
    cell_b.font = BODY_FONT
    cell_c = ws.cell(row=row, column=3, value=value)
    cell_c.font = BODY_BOLD
    if fmt:
        cell_c.number_format = fmt
    fill = ALT_FILL if alt else WHITE_FILL
    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = fill
    return row + 1


def style_data_row(ws, row, max_col, alt=False):
    """Apply alternating row fill."""
    fill = ALT_FILL if alt else WHITE_FILL
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill


def _write_scenario_block(ws, row, scenario_label, params, results, aggregated,
                          n_months, max_col, is_first=False):
    """Write one scenario's full output block. Returns the next row."""

    # ── Scenario Title ────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=scenario_label)
    cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # ── Calibration Parameters ────────────────────────────────
    row = write_section_header(ws, row, "CALIBRATION PARAMETERS", max_col)
    row += 1

    for regime_key, regime_label in [("pre_halving", "Months 1–24 (Pre-Halving)"),
                                     ("post_halving", "Months 25–60 (Post-Halving)")]:
        p = params[regime_key]
        ws.cell(row=row, column=2, value=regime_label).font = SECTION_FONT
        row += 1
        row = write_kv_row(ws, row, "Price mean log-return (monthly)", p["mu_price"], FMT_NUMBER_6)
        row = write_kv_row(ws, row, "Price std dev (monthly)", p["sigma_price"], FMT_NUMBER_6, alt=True)
        row = write_kv_row(ws, row, "Difficulty mean log-return (monthly)", p["mu_diff"], FMT_NUMBER_6)
        row = write_kv_row(ws, row, "Difficulty std dev (monthly)", p["sigma_diff"], FMT_NUMBER_6, alt=True)
        row = write_kv_row(ws, row, "Price-Difficulty correlation", p["correlation"], FMT_NUMBER_DEC)
        row += 1

    row += 1

    # ── Summary Statistics ────────────────────────────────────
    row = write_section_header(ws, row, "SUMMARY STATISTICS & PROBABILITY TABLE", max_col)
    row += 1

    stats = aggregated["stats"]
    probs = aggregated["probs"]

    pct_headers = ["", "Metric", "P10", "P25", "P50 (Median)", "P75", "P90", "Mean"]
    row = write_sub_header(ws, row, pct_headers, max_col)

    metrics_rows = [
        ("IRR", "irr", FMT_PCT),
        ("MOIC", "moic", FMT_NUMBER_DEC),
        ("Payback Month", "payback", FMT_NUMBER),
        ("Year-1 Net Income", "y1_net_income", FMT_CURRENCY),
        ("5-Year Net Income", "y5_net_income", FMT_CURRENCY),
    ]
    for idx, (label, key, fmt) in enumerate(metrics_rows):
        alt = idx % 2 == 1
        ws.cell(row=row, column=2, value=label).font = BODY_BOLD
        for j, p in enumerate([10, 25, 50, 75, 90]):
            cell = ws.cell(row=row, column=3 + j, value=stats[key][f"P{p}"])
            cell.number_format = fmt
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
        cell = ws.cell(row=row, column=8, value=stats[key]["mean"])
        cell.number_format = fmt
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="center")
        style_data_row(ws, row, max_col, alt)
        row += 1

    row += 1

    ws.cell(row=row, column=2, value="Probability Thresholds").font = SECTION_FONT
    row += 1
    prob_items = [
        ("Prob. IRR > 0%", probs["irr_gt_0"]),
        ("Prob. IRR > 10%", probs["irr_gt_10"]),
        ("Prob. IRR > 20%", probs["irr_gt_20"]),
        ("Prob. Payback within 24 months", probs["payback_24"]),
        ("Prob. Payback within 36 months", probs["payback_36"]),
        ("Prob. Payback within 48 months", probs["payback_48"]),
        ("Prob. Losing Money (5-yr)", probs["lose_money"]),
    ]
    for i, (label, val) in enumerate(prob_items):
        row = write_kv_row(ws, row, label, val, FMT_PCT, alt=(i % 2 == 1))

    row += 2

    # ── Cumulative FCF Percentiles ────────────────────────────
    row = write_section_header(ws, row, "CUMULATIVE FREE CASHFLOW PERCENTILES (MONTH-BY-MONTH)", max_col)
    row += 1

    cf_headers = ["", "Month", "Date", "P10", "P25", "P50", "P75", "P90"]
    row = write_sub_header(ws, row, cf_headers, max_col)

    schedule = results["schedule"]
    cf_pct = aggregated["cf_percentiles"]
    n_pts = len(cf_pct["P10"])

    # Month 0
    ws.cell(row=row, column=2, value=0).font = BODY_FONT
    ws.cell(row=row, column=3, value="Mar-2026").font = BODY_FONT
    for j, p_key in enumerate(["P10", "P25", "P50", "P75", "P90"]):
        cell = ws.cell(row=row, column=4 + j, value=cf_pct[p_key][0])
        cell.number_format = FMT_CURRENCY
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="right")
    style_data_row(ws, row, max_col, alt=False)
    row += 1

    for m_idx in range(min(n_months, n_pts - 1)):
        yr, mo, dim, dt = schedule[m_idx]
        alt = (m_idx + 1) % 2 == 1
        ws.cell(row=row, column=2, value=m_idx + 1).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=dt.strftime("%b-%Y")).font = BODY_FONT
        for j, p_key in enumerate(["P10", "P25", "P50", "P75", "P90"]):
            cell = ws.cell(row=row, column=4 + j, value=cf_pct[p_key][m_idx + 1])
            cell.number_format = FMT_CURRENCY
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="right")
        style_data_row(ws, row, max_col, alt)
        row += 1

    row += 2

    # ── IRR Distribution ──────────────────────────────────────
    row = write_section_header(ws, row, "IRR DISTRIBUTION (HISTOGRAM)", max_col)
    row += 1

    hist_headers = ["", "Bucket Lower", "Bucket Upper", "Count", "Frequency"]
    row = write_sub_header(ws, row, hist_headers, max_col)

    hist_counts = aggregated["hist_counts"]
    hist_edges = aggregated["hist_edges"]
    total_count = hist_counts.sum() if len(hist_counts) > 0 else 1

    for i in range(len(hist_counts)):
        alt = i % 2 == 1
        ws.cell(row=row, column=2, value=hist_edges[i]).font = BODY_FONT
        ws.cell(row=row, column=2).number_format = FMT_PCT
        ws.cell(row=row, column=3, value=hist_edges[i + 1]).font = BODY_FONT
        ws.cell(row=row, column=3).number_format = FMT_PCT
        ws.cell(row=row, column=4, value=int(hist_counts[i])).font = BODY_FONT
        ws.cell(row=row, column=4).number_format = FMT_NUMBER
        freq = hist_counts[i] / total_count
        ws.cell(row=row, column=5, value=freq).font = BODY_FONT
        ws.cell(row=row, column=5).number_format = FMT_PCT2
        style_data_row(ws, row, max_col, alt)
        row += 1

    row += 2

    # ── Sample Paths ──────────────────────────────────────────
    row = write_section_header(ws, row, "SAMPLE SIMULATION PATHS (20 PATHS)", max_col)
    row += 1

    sample_paths = results["sample_paths"]
    path_headers = ["", "Month", "Date", "BTC Price", "Difficulty", "Hashprice",
                    "Revenue", "Net Income", "Cum. Cashflow"]

    for path_idx, sim_id in enumerate(sorted(sample_paths.keys())):
        path = sample_paths[sim_id]
        ws.cell(row=row, column=2, value=f"Path #{path_idx + 1} (sim {sim_id})").font = SECTION_FONT
        row += 1
        row = write_sub_header(ws, row, path_headers, max_col)

        cum_cf = path["cum_cf"]
        for m_idx, pd_row in enumerate(path["path_data"]):
            alt = m_idx % 2 == 1
            ws.cell(row=row, column=2, value=pd_row["month"]).font = BODY_FONT
            ws.cell(row=row, column=3, value=pd_row["date"].strftime("%b-%Y")).font = BODY_FONT
            ws.cell(row=row, column=4, value=pd_row["btc_price"]).font = BODY_FONT
            ws.cell(row=row, column=4).number_format = FMT_CURRENCY
            ws.cell(row=row, column=5, value=pd_row["difficulty"]).font = BODY_FONT
            ws.cell(row=row, column=5).number_format = "0.00E+00"
            ws.cell(row=row, column=6, value=pd_row["hashprice"]).font = BODY_FONT
            ws.cell(row=row, column=6).number_format = FMT_CURRENCY_DEC
            ws.cell(row=row, column=7, value=pd_row["revenue"]).font = BODY_FONT
            ws.cell(row=row, column=7).number_format = FMT_CURRENCY
            ws.cell(row=row, column=8, value=pd_row["net_income"]).font = BODY_FONT
            ws.cell(row=row, column=8).number_format = FMT_CURRENCY
            ws.cell(row=row, column=9, value=cum_cf[m_idx + 1]).font = BODY_FONT
            ws.cell(row=row, column=9).number_format = FMT_CURRENCY
            style_data_row(ws, row, max_col, alt)
            row += 1

        row += 1

    return row


def write_mc_sheet(wb, params, results, aggregated, n_months):
    """Create the MC Simulation sheet in the workbook."""
    print("\n" + "=" * 60)
    print("STEP 5: WRITING EXCEL SHEET")
    print("=" * 60)

    if "MC Simulation" in wb.sheetnames:
        del wb["MC Simulation"]

    ws = wb.create_sheet("MC Simulation")
    MAX_COL = 10
    row = 1

    # ── Title ─────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    cell = ws.cell(row=row, column=1,
                   value="MONTE CARLO SIMULATION: Project 1 (Omerelu 10MW)")
    cell.font = TITLE_FONT
    for c in range(1, MAX_COL + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    cell = ws.cell(row=row, column=1,
                   value=f"Generated: {ts}  |  10,000 simulations x 60 months  |  Seed: 42  |  Operational shutdown floor applied")
    cell.font = Font(name="Calibri", size=9, italic=True, color="808080")
    row += 2

    # ── Fixed Project Parameters ──────────────────────────────
    row = write_section_header(ws, row, "FIXED PROJECT PARAMETERS", MAX_COL)
    row += 1
    fixed_params = [
        ("Starting BTC Price", STARTING_BTC_PRICE, FMT_CURRENCY),
        ("Starting Difficulty", STARTING_DIFFICULTY, "#,##0"),
        ("Capacity", f"{CAPACITY_MW} MW", None),
        ("Fleet Hashrate", f"{FLEET_HASHRATE_THS:,} TH/s", None),
        ("Uptime", UPTIME, FMT_PCT),
        ("PPA Rate", PPA_RATE, FMT_CURRENCY_DEC),
        ("O&M Rate", OM_RATE, FMT_CURRENCY_DEC),
        ("Axxela Revenue Share", AXXELA_SHARE, FMT_PCT),
        ("Tax Rate", TAX_RATE, FMT_PCT),
        ("Monthly Depreciation", DEPRECIATION_MONTHLY, FMT_CURRENCY_DEC),
        ("Total CapEx", TOTAL_CAPEX, FMT_CURRENCY_DEC),
        ("Halving Difficulty Shock", DIFFICULTY_SHOCK, FMT_PCT),
        ("Operational Shutdown", "Mine idles when EBITDA < 0 (FCF floored at $0)", None),
    ]
    for i, (label, val, fmt) in enumerate(fixed_params):
        row = write_kv_row(ws, row, label, val, fmt, alt=(i % 2 == 1))
    row += 2

    # ── Write single scenario block ───────────────────────────
    row = _write_scenario_block(
        ws, row,
        "MONTE CARLO RESULTS",
        params, results, aggregated,
        n_months, MAX_COL,
    )

    # Column widths
    col_widths = {1: 4, 2: 40, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18, 9: 18, 10: 4}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    print("  MC Simulation sheet written successfully.")
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Monte Carlo Mining Simulation")
    parser.add_argument("--simulations", type=int, default=10_000, help="Number of simulations")
    parser.add_argument("--months", type=int, default=60, help="Projection months")
    parser.add_argument("--data-path", type=str, default="data/btc_merged_data.xlsx",
                        help="Path to historical data Excel file")
    parser.add_argument("--model-path", type=str,
                        default="AGM Consolidated Model AA_2.0.xlsx",
                        help="Path to consolidated model Excel file")
    parser.add_argument("--output-dir", type=str, default="output",
                        help="Output directory")
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  MONTE CARLO BITCOIN MINING SIMULATION")
    print("  AGM Omerelu 10MW Project")
    print("=" * 60)
    print(f"  Simulations: {args.simulations:,}")
    print(f"  Months: {args.months}")
    print(f"  Data: {args.data_path}")
    print(f"  Model: {args.model_path}")
    print(f"  Output: {args.output_dir}/")

    # Ensure output dir exists
    os.makedirs(args.output_dir, exist_ok=True)

    # Step 1: Calibrate
    all_params = calibrate(args.data_path)
    params = all_params["bull"]

    # Step 2 & 3: Simulate
    results = run_simulations(params, args.simulations, args.months)

    # Step 4: Aggregate
    aggregated = aggregate(results)

    # Step 5: Write Excel
    output_file = os.path.join(args.output_dir, "AGM_Consolidated_Model_AA_2_0.xlsx")
    print(f"\n  Copying model to {output_file} ...")
    shutil.copy2(args.model_path, output_file)

    wb = load_workbook(output_file)
    write_mc_sheet(wb, params, results, aggregated, args.months)
    wb.save(output_file)
    print(f"\n  Output saved: {output_file}")
    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)


if __name__ == "__main__":
    main()
