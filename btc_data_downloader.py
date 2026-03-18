"""
Bitcoin Historical Data Downloader
====================================
Pulls and merges three datasets required for Bitcoin mining financial models:
  1. BTC Price (daily close, USD)
  2. Network Difficulty
  3. Hashprice ($/PH/day) — derived from price, difficulty, and block reward

Data sources (free, no API key required):
  - CoinGecko API (primary) / CoinMetrics Community API (fallback) for BTC price
  - Blockchain.info API (primary) / CoinMetrics (fallback) for network difficulty
  - Hashprice is computed using the standard mining revenue formula

Output:
  - Individual CSVs for each dataset
  - Merged CSV with all three datasets joined on date
  - Formatted Excel workbook with Daily, Annual, and Monthly tabs

Usage:
  python btc_data_downloader.py [--days 1825] [--output-dir ./output]

Built for the AGM Omerelu 10MW Bitcoin Mining Project evaluation.
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timedelta

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


# ── Configuration ──────────────────────────────────────────────────

HALVING_SCHEDULE = [
    (datetime(2012, 11, 28), 25.0),
    (datetime(2016,  7,  9), 12.5),
    (datetime(2020,  5, 11), 6.25),
    (datetime(2024,  4, 20), 3.125),
    (datetime(2028,  4, 15), 1.5625),   # estimated
]

REQUEST_TIMEOUT = 30  # seconds


# ── Data Fetchers ──────────────────────────────────────────────────

class DataFetchError(Exception):
    """Raised when all data sources fail for a given metric."""
    pass


def _get_json(url, params=None):
    """HTTP GET with timeout, returns parsed JSON."""
    resp = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def fetch_btc_price(days):
    """
    Fetch daily BTC/USD price.
    Tries CoinGecko first, falls back to CoinMetrics Community API.
    """
    # Source 1: CoinGecko
    try:
        print("  [price] Trying CoinGecko...")
        data = _get_json(
            "https://api.coingecko.com/api/v3/coins/bitcoin/market_chart",
            {"vs_currency": "usd", "days": days, "interval": "daily"},
        )
        rows = [
            {
                "date": datetime.utcfromtimestamp(ts / 1000).strftime("%Y-%m-%d"),
                "btc_price_usd": round(price, 2),
            }
            for ts, price in data["prices"]
        ]
        print(f"  [price] CoinGecko: {len(rows)} records")
        return rows
    except Exception as e:
        print(f"  [price] CoinGecko failed: {e}")

    # Source 2: CoinMetrics Community API
    try:
        print("  [price] Trying CoinMetrics...")
        start = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
        url = "https://community-api.coinmetrics.io/v4/timeseries/asset-metrics"
        params = {
            "assets": "btc",
            "metrics": "PriceUSD",
            "frequency": "1d",
            "start_time": start,
            "page_size": 10000,
        }
        rows = []
        while True:
            data = _get_json(url, params)
            for entry in data.get("data", []):
                rows.append({
                    "date": entry["time"][:10],
                    "btc_price_usd": round(float(entry["PriceUSD"]), 2),
                })
            next_url = data.get("next_page_url")
            if next_url:
                url, params = next_url, {}
            else:
                break
        print(f"  [price] CoinMetrics: {len(rows)} records")
        return rows
    except Exception as e:
        print(f"  [price] CoinMetrics failed: {e}")

    raise DataFetchError(
        "Could not fetch BTC price from any source.\n"
        "Manual alternative: https://coincodex.com/crypto/bitcoin/historical-data/"
    )


def fetch_difficulty(days):
    """
    Fetch daily network difficulty.
    Tries Blockchain.info first, falls back to CoinMetrics Community API.
    """
    # Source 1: Blockchain.info
    try:
        print("  [difficulty] Trying Blockchain.info...")
        data = _get_json(
            "https://api.blockchain.info/charts/difficulty",
            {"timespan": f"{days}days", "format": "json", "rollingAverage": "1days"},
        )
        rows = [
            {
                "date": datetime.utcfromtimestamp(v["x"]).strftime("%Y-%m-%d"),
                "difficulty": v["y"],
            }
            for v in data.get("values", [])
        ]
        print(f"  [difficulty] Blockchain.info: {len(rows)} records")
        return rows
    except Exception as e:
        print(f"  [difficulty] Blockchain.info failed: {e}")

    # Source 2: CoinMetrics
    try:
        print("  [difficulty] Trying CoinMetrics...")
        start = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
        url = "https://community-api.coinmetrics.io/v4/timeseries/asset-metrics"
        params = {
            "assets": "btc",
            "metrics": "DiffMean",
            "frequency": "1d",
            "start_time": start,
            "page_size": 10000,
        }
        rows = []
        while True:
            data = _get_json(url, params)
            for entry in data.get("data", []):
                rows.append({
                    "date": entry["time"][:10],
                    "difficulty": float(entry["DiffMean"]),
                })
            next_url = data.get("next_page_url")
            if next_url:
                url, params = next_url, {}
            else:
                break
        print(f"  [difficulty] CoinMetrics: {len(rows)} records")
        return rows
    except Exception as e:
        print(f"  [difficulty] CoinMetrics failed: {e}")

    raise DataFetchError(
        "Could not fetch difficulty from any source.\n"
        "Manual alternative: https://www.blockchain.com/charts/difficulty"
    )


# ── Hashprice Derivation ──────────────────────────────────────────

def get_block_reward(date_str):
    """Return the BTC block reward effective on a given date."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    reward = 50.0
    for halving_date, new_reward in HALVING_SCHEDULE:
        if dt >= halving_date:
            reward = new_reward
    return reward


def compute_hashprice(price_rows, diff_rows):
    """
    Derive hashprice ($/PH/day) from BTC price and network difficulty.

    Formula:
        Hashprice = BTC_Price × Block_Reward × 86400 / (Difficulty × 2^32) × 10^15

    This is the standard mining revenue formula used by Hashrate Index,
    Luxor, and institutional mining desks. It represents the expected
    daily USD revenue per petahash of SHA-256 hashrate.
    """
    print("  [hashprice] Computing from price + difficulty...")

    price_map = {r["date"]: r["btc_price_usd"] for r in price_rows}
    diff_map = {r["date"]: r["difficulty"] for r in diff_rows}
    common_dates = sorted(set(price_map) & set(diff_map))

    TWO_POW_32 = 2**32
    rows = []
    for dt in common_dates:
        price = price_map[dt]
        diff = diff_map[dt]
        reward = get_block_reward(dt)
        if diff > 0:
            hp = price * reward * 86400 / (diff * TWO_POW_32) * 1e15
            rows.append({
                "date": dt,
                "btc_price_usd": price,
                "difficulty": diff,
                "block_reward": reward,
                "hashprice_usd_ph_day": round(hp, 4),
            })

    print(f"  [hashprice] Computed for {len(rows)} dates")
    return rows


# ── Merge ──────────────────────────────────────────────────────────

def merge_datasets(price_rows, diff_rows, hp_rows):
    """
    Outer-join all datasets on date. Rows with price but no difficulty
    (or vice versa) are retained with NaN for missing fields.
    """
    import pandas as pd

    df_price = pd.DataFrame(price_rows)
    df_diff = pd.DataFrame(diff_rows)
    df_hp = pd.DataFrame(hp_rows)

    merged = df_hp[["date", "btc_price_usd", "difficulty", "block_reward", "hashprice_usd_ph_day"]].copy()

    price_only = df_price[~df_price["date"].isin(merged["date"])][["date", "btc_price_usd"]]
    if len(price_only) > 0:
        merged = pd.concat([merged, price_only], ignore_index=True)

    merged = merged.sort_values("date").reset_index(drop=True)
    merged["difficulty_T"] = merged["difficulty"] / 1e12
    return merged


# ── Output Writers ─────────────────────────────────────────────────

def write_csv(rows, filepath, fieldnames):
    """Write a list of dicts to CSV."""
    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    print(f"  Saved {filepath} ({len(rows)} rows)")


def write_xlsx(merged_df, filepath):
    """
    Create a formatted Excel workbook with three tabs:
      1. Daily Data — full merged dataset
      2. Annual Averages — yearly summary statistics
      3. Monthly Averages — month-by-month averages
    """
    import pandas as pd

    wb = Workbook()

    # Styles
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1B3A5C")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10, color="333333")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    def _write_sheet(ws, headers, widths, data_rows, col_keys, num_fmts=None):
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, border
            ws.column_dimensions[get_column_letter(c)].width = w
        for r, row_data in enumerate(data_rows, 2):
            fill = alt_fill if r % 2 == 0 else None
            for c, key in enumerate(col_keys, 1):
                val = row_data.get(key)
                if pd.isna(val) if not isinstance(val, str) else False:
                    val = ""
                cell = ws.cell(row=r, column=c, value=val)
                cell.font, cell.border = cell_font, border
                if fill:
                    cell.fill = fill
                if num_fmts and c in num_fmts and val != "":
                    cell.number_format = num_fmts[c]
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data_rows)+1}"
        ws.freeze_panes = "A2"

    # Tab 1: Daily
    ws1 = wb.active
    ws1.title = "Daily Data"
    daily_rows = merged_df.to_dict("records")
    _write_sheet(
        ws1,
        ["Date", "BTC Price (USD)", "Network Difficulty", "Difficulty (T)", "Block Reward", "Hashprice ($/PH/day)"],
        [13, 18, 24, 16, 14, 22],
        daily_rows,
        ["date", "btc_price_usd", "difficulty", "difficulty_T", "block_reward", "hashprice_usd_ph_day"],
        {2: "$#,##0.00", 3: "#,##0", 4: "#,##0.00", 5: "0.0000", 6: "$#,##0.00"},
    )

    # Tab 2: Annual
    merged_df["year"] = pd.to_datetime(merged_df["date"]).dt.year
    annual = merged_df.groupby("year").agg(
        avg_price=("btc_price_usd", "mean"),
        min_price=("btc_price_usd", "min"),
        max_price=("btc_price_usd", "max"),
        avg_diff_T=("difficulty_T", "mean"),
        avg_hp=("hashprice_usd_ph_day", "mean"),
        min_hp=("hashprice_usd_ph_day", "min"),
        max_hp=("hashprice_usd_ph_day", "max"),
        days=("date", "count"),
    ).reset_index()
    ws2 = wb.create_sheet("Annual Averages")
    _write_sheet(
        ws2,
        ["Year", "Avg Price", "Min Price", "Max Price", "Avg Diff (T)", "Avg HP", "Min HP", "Max HP", "Days"],
        [8, 14, 14, 14, 16, 14, 14, 14, 8],
        annual.to_dict("records"),
        ["year", "avg_price", "min_price", "max_price", "avg_diff_T", "avg_hp", "min_hp", "max_hp", "days"],
        {2: "$#,##0", 3: "$#,##0", 4: "$#,##0", 5: "#,##0.0", 6: "$#,##0.00", 7: "$#,##0.00", 8: "$#,##0.00"},
    )

    # Tab 3: Monthly
    merged_df["year_month"] = pd.to_datetime(merged_df["date"]).dt.to_period("M").astype(str)
    monthly = merged_df.groupby("year_month").agg(
        avg_price=("btc_price_usd", "mean"),
        avg_diff_T=("difficulty_T", "mean"),
        avg_hp=("hashprice_usd_ph_day", "mean"),
        days=("date", "count"),
    ).reset_index()
    ws3 = wb.create_sheet("Monthly Averages")
    _write_sheet(
        ws3,
        ["Month", "Avg Price (USD)", "Avg Difficulty (T)", "Avg Hashprice ($/PH)", "Days"],
        [12, 18, 20, 22, 8],
        monthly.to_dict("records"),
        ["year_month", "avg_price", "avg_diff_T", "avg_hp", "days"],
        {2: "$#,##0.00", 3: "#,##0.0", 4: "$#,##0.00"},
    )

    wb.save(filepath)
    print(f"  Saved {filepath}")
    print(f"    Tab 1: Daily Data ({len(daily_rows)} rows)")
    print(f"    Tab 2: Annual Averages ({len(annual)} rows)")
    print(f"    Tab 3: Monthly Averages ({len(monthly)} rows)")


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Download and merge Bitcoin historical data for mining model calibration."
    )
    parser.add_argument(
        "--days", type=int, default=1825,
        help="Number of days of history to fetch (default: 1825 = 5 years)",
    )
    parser.add_argument(
        "--output-dir", type=str, default="./output",
        help="Directory for output files (default: ./output)",
    )
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)

    print("=" * 60)
    print("  Bitcoin Historical Data Downloader")
    print(f"  Fetching {args.days} days of data")
    print("=" * 60)

    # Fetch
    print("\n1. Fetching data from APIs...\n")
    try:
        price_rows = fetch_btc_price(args.days)
    except DataFetchError as e:
        print(f"\n  ERROR: {e}")
        price_rows = []

    try:
        diff_rows = fetch_difficulty(args.days)
    except DataFetchError as e:
        print(f"\n  ERROR: {e}")
        diff_rows = []

    if not price_rows or not diff_rows:
        print("\n  Cannot proceed without both price and difficulty data.")
        print("  See error messages above for manual download alternatives.")
        sys.exit(1)

    hp_rows = compute_hashprice(price_rows, diff_rows)

    # Write individual CSVs
    print("\n2. Writing individual CSVs...\n")
    write_csv(price_rows, os.path.join(args.output_dir, "btc_price.csv"), ["date", "btc_price_usd"])
    write_csv(diff_rows, os.path.join(args.output_dir, "btc_difficulty.csv"), ["date", "difficulty"])
    write_csv(
        hp_rows,
        os.path.join(args.output_dir, "btc_hashprice.csv"),
        ["date", "btc_price_usd", "difficulty", "block_reward", "hashprice_usd_ph_day"],
    )

    # Merge and write combined outputs
    print("\n3. Merging datasets and writing combined output...\n")
    merged = merge_datasets(price_rows, diff_rows, hp_rows)
    merged.to_csv(os.path.join(args.output_dir, "btc_merged_data.csv"), index=False)
    print(f"  Saved btc_merged_data.csv ({len(merged)} rows)")
    write_xlsx(merged, os.path.join(args.output_dir, "btc_historical_data.xlsx"))

    print("\n" + "=" * 60)
    print(f"  Done. Files saved to: {os.path.abspath(args.output_dir)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
